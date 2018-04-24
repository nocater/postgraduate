import openpyxl
import numpy as np
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import OneHotEncoder
from sklearn.utils import shuffle


def getdata(file):
    data = []
    # data_only=True 保证取到的公式是数值
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[0]]

    #取出所有数据
    cell_range = ws['B4':'AE250']
    cell_range = ws['B5':'Z596']
    for row in cell_range:
        arr = [cell.value if cell.value is not None else 0. for cell in row]
        data.append(arr)
    # 校验
    data = np.array(data)
    # assert np.dtype == float

    #删除无用维度 选取所有成分 等温温度 回火温度 强度(Y)
    all_dimension = set(i for i in range(32))
    need_dimension = set(i for i in range(7)) | {12, 15, 24}
    data = np.delete(data, list(all_dimension - need_dimension), axis=1)
    data = data.astype(float)

    # 删除强度为0的行数据
    retains = data[:, -1].nonzero()[0]
    data = data[retains, :]
    data = shuffle(data)

    # 回归变成分类问题
    x = data[:, :-1]
    y = data[:, -1:]
    y = deal_labels(y)

    # 数据缩放 way 2 x_norm = np.linalg.norm(x, axis = 1, keepdims = True) x = x/x_norm
    scaler = MinMaxScaler()
    x = scaler.fit_transform(x)
    """"""
    # 皮尔斯恩系数
    # pearson = []
    # print(x.dtype, y.dtype)
    # for i in range(x.shape[1]):
    #     x_ = np.hstack(x[:, i])
    #     y_ = np.hstack(y)
    #     pearson.append(np.corrcoef(x_, y_)[1, 0])
    # print(pearson)
    # y = deal_labels(y)

    print('数据维度: x', x.shape, ' y', y.shape, sep='')
    return x, y


def deal_labels(y, categories=4):
    """
    将向量分类表示
    """
    # assert '<' not in str(y.dtype)
    y = y.astype(float)
    min_ = np.min(y)
    max_ = np.max(y)
    # 设置分类边界
    bounds = np.linspace(min_-1, max_, categories+1)
    print('类别统计:', np.histogram(y, bounds)[0])
    # 转换类别编号
    y = [(bounds >= i).nonzero()[0][0] for i in y]
    # 编号下标从0开始
    y = y-np.array([1])
    y = y.reshape(len(list(y)), 1)
    enc = OneHotEncoder()
    y = enc.fit_transform(y).toarray()
    return y

if __name__ == '__main__':
    dir = r'C:\Users\chenshuai\Documents\材料学院\贝氏体钢数据统计-chenshuai.xlsx'
    dir2 = r'C:\Users\chenshuai\Documents\材料学院\贝氏体钢数据统计-总20180410.xlsx'
    x, y = getdata(dir)
    # *chenshuai.xlsx 数据预处理完成 原式文件并没有(<100) 完成了数据预处理代码